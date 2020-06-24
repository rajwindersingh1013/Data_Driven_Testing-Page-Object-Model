

from Src.Base.Setup import EnvironmentSetup
import openpyxl


class Excel_Reader():
	def __init__(self,sheetname):
		# set file path
		global filepath
		__sheetname =sheetname
		__filepath = "F:\Pyhton\Car_Registration\Test\TestData\CarList.xlsx"


		try:
			self.__workbook =openpyxl.load_workbook(__filepath)
			self.__sheet =self.__workbook.get_sheet_by_name(__sheetname)

		except:
			print("Sheet is not located")



	def get_number_of_columns(self):
		return self.__sheet.max_column

	def get_number_of_rows(self):
		return self.__sheet.max_row




	def get_data_Cellwise_full_table(self):
		rows=self.get_number_of_rows()

		cols=self.get_number_of_columns()
		print("rows", rows, "  columns", cols)
		data =[[0]*cols]*(rows-1)
		for r in range(rows-1):
			for c in range(cols):
				print(c)
				data[r][c]=self.__sheet.cell(row=r+2,column=7).value
				print(data[r][c],end=" ")
			print(r)
		return data


	def getfulldata(self):
		data=[]
		cell=[]
		for row in self.__sheet.values:
			data.append(row)

		data.remove(data[0])
		return(data)






	def get_data_row_wise(self,rownum):
		for c in range(1, self.get_number_of_columns() + 1):
			return(self.__sheet.cell(row=rownum, column=c).value)

	def get_data_columnwise(self,colnum,torow):
		data=[[]*(torow-1)]
		for r in range(torow-1):
			data[r]=self.__sheet.cell(row=r+1,column =colnum).value
		return data



